import pandas as pd
import win32com.client
from PIL import ImageGrab
import openpyxl
from openpyxl.utils import get_column_letter
import os
from datetime import datetime
import time 

class AutomacaoEmailOutlook:
    
    def __init__(self, arquivo_destinatarios, arquivo_dados, aba_destinatarios):
        """
        arquivo_destinatarios: Planilha com os emails
        arquivo_dados: Planilha com os dados E-REP
        aba_destinatarios: Nome da aba onde estão os emails
        """
        self.arquivo_destinatarios = arquivo_destinatarios
        self.arquivo_dados = arquivo_dados
        self.aba_destinatarios = aba_destinatarios 
        self.outlook = win32com.client.Dispatch('outlook.application')
        
        # --- ATUALIZE AQUI O NOME DO ARQUIVO DA IMAGEM DA SUA ASSINATURA ---
        # (Esta imagem deve estar na mesma pasta do script)
        self.arquivo_imagem_assinatura = "logo_assinatura.png" # <--- MUDE AQUI
        # ------------------------------------------------------------------

    def _gerar_html_assinatura(self, cid_imagem_assinatura):
        """
        Gera o HTML da assinatura personalizada com negrito seletivo.
        """
        
        texto_assinatura_html = f"""
        <div style="font-family: Aptos, Arial, sans-serif; font-size: 12pt; font-weight: normal;">

            <br>
            <p style="font-style: italic; margin-top: 15px; font-weight: bold;">
                "Na Renault, a Qualidade não é opcional e sim, um diferencial competitivo. Promover a qualidade é fazer TODA a diferença para o Cliente."
            </p>
            
            <p style="margin-top: 15px; margin-bottom: 15px;">
                Atenciosamente,
            </p>
            
            <table cellpadding="0" cellspacing="0" border="0" style="font-family: Aptos, Arial, sans-serif; font-size: 12pt; font-weight: normal; color: #222; line-height: 1.4;">
                <tr>
                    <td style="padding-right: 15px; vertical-align: top;">
                        <img src="cid:{cid_imagem_assinatura}">
                    </td>
                    
                    <td style="vertical-align: top;">
                        <p style="margin: 0; color: #000;"><strong>
                            RODRIGO DE SOUSA
                        </strong></p>
                        <p style="margin: 0;">
                            QMR | DIREÇÃO PÓS-VENDAS
                        </p>
                        <p style="margin: 0;">
                            QUALIDADE E MÉTODOS REDE
                        </p>
                        <p style="margin: 0;">
                            rodrigo.sousa-extern@renault.com
                        </p>
                        <p style="margin: 0;">
                            (11) 94290-6904
                        </p>
                    </td>
                </tr>
            </table>
        </div>
        """
        
        return texto_assinatura_html

    def carregar_destinatarios(self):
        """Carrega a planilha de destinatários e limpa os nomes dos grupos"""
        try:
            df = pd.read_excel(self.arquivo_destinatarios, sheet_name=self.aba_destinatarios)
        except Exception as e:
            print(f"✗ ERRO AO CARREGAR DESTINATÁRIOS de '{self.arquivo_destinatarios}' (Aba: {self.aba_destinatarios}): {e}")
            print("  Verifique o nome do arquivo, a extensão (.xlsm) e o nome da aba.")
            return None
            
        if 'Grupo' in df.columns:
            # Limpa "Grupo R Point" para ficar apenas "R Point"
            df['Grupo'] = df['Grupo'].astype(str).str.replace("Grupo ", "").str.strip()
        else:
            print(f"✗ ERRO: Coluna 'Grupo' não encontrada no arquivo de destinatários.")
            return None
            
        return df.groupby('Grupo').agg({
            'Para': lambda x: '; '.join(x.dropna()),
            'Cc': lambda x: '; '.join(x.dropna()) if x.notna().any() else ''
        }).reset_index()
    
    def tirar_screenshot_tabela(self, arquivo_excel, aba, linha_inicio, linha_fim, 
                                coluna_inicio, coluna_fim, caminho_saida):
        """Tira screenshot de uma tabela específica no Excel"""
        
        # Define as variáveis ANTES do try para que o 'finally' possa acessá-las
        excel = None
        wb = None
        
        try:
            excel = win32com.client.Dispatch('Excel.Application')
            excel.Visible = False
            excel.DisplayAlerts = False
        
            wb = excel.Workbooks.Open(os.path.abspath(arquivo_excel))
            ws = wb.Sheets(aba)
            col_start = get_column_letter(coluna_inicio)
            col_end = get_column_letter(coluna_fim)
            intervalo = f"{col_start}{linha_inicio}:{col_end}{linha_fim}"
            
            ws.Range(intervalo).CopyPicture(Appearance=1, Format=2)
            time.sleep(0.5) 
            
            img = ImageGrab.grabclipboard()
            if img:
                img.save(caminho_saida)
                print(f"  ✓ Screenshot salvo: {caminho_saida}")
                return caminho_saida
            else:
                print(f"  ✗ Clipboard estava vazio. Não foi possível salvar o screenshot.")
                return None
                
        except Exception as e:
            print(f"  ✗ Erro ao tirar screenshot: {e}")
            return None
        finally:
            # Garante que o processo do Excel seja finalizado
            if wb:
                wb.Close(SaveChanges=False)
                del wb  # Força a liberação do objeto
                
            if excel:
                excel.Quit()
                del excel # Força a liberação do objeto
    
    def gerar_mensagem_alerta(self, nome_grupo, lista_desenquadradas, cid_imagem):
        """Gera a mensagem de ALERTA do email com o novo template e fonte"""
        
        # --- Lógica para formatar a lista de concessionárias ---
        nomes_formatados_html = [f'<strong class="alerta">{nome}</strong>' for nome in lista_desenquadradas]
        
        texto_principal = ""
        if len(nomes_formatados_html) == 1:
            str_nomes = nomes_formatados_html[0]
            texto_principal = f"A concessionária {str_nomes} está desenquadrada."
        elif len(nomes_formatados_html) == 2:
            str_nomes = f"{nomes_formatados_html[0]} e {nomes_formatados_html[1]}"
            texto_principal = f"As concessionárias {str_nomes} estão desenquadradas."
        elif len(nomes_formatados_html) > 2:
            # Junta todos, exceto o último, com vírgula. Adiciona o último com "e".
            str_nomes = f"{', '.join(nomes_formatados_html[:-1])}, e {nomes_formatados_html[-1]}"
            texto_principal = f"As concessionárias {str_nomes} estão desenquadradas."
        # --- Fim da lógica de formatação ---

        mensagem = f"""
        <html>
        <head>
            <style>
                body {{ 
                    font-family: Aptos, Arial, sans-serif;
                    font-size: 12pt;
                    font-weight: normal;
                }}
                .alerta {{ 
                    color: #D32F2F;
                    font-weight: bold; 
                }}
            </style>
        </head>
        <body>
            <p>Boa tarde a todos!</p>
            <p>Espero que estejam bem.</p>
            
            <p>Segue o acompanhamento do GMB do Grupo <strong>{nome_grupo}</strong>:</p>
            
            <p class="alerta">{texto_principal}</p>
            
            <p>Utilizar as <strong>dicas disponíveis
            no portal Reputation</strong> e <strong>seguir rigorosamente as boas práticas</strong>
            recomendadas contribuirá significativamente para a melhoria dos resultados. <strong class="alerta"> A negligência quanto a esses passos poderá agravar
            mais o cenário atual.</strong></p>
            
            <hr>
            <p style="font-style: italic;">Detalhes da apuração:</p>
            <p><img src="cid:{cid_imagem}"></p>

            <p><strong><i>É fundamental que todos intensifiquem os esforços para
            mudar esse quadro.</i></strong> Com o engajamento de todos, será possível reverter a
            situação atual e alcançar o objetivo desejado.</p>
            
            <p>Contamos com a colaboração e empenho de vocês para reverter
            a situação o mais rápido possível.</p>
            
        </body>
        </html>
        """
        return mensagem
        
    def gerar_mensagem_parabens(self, concessionaria, cid_imagem):
        """Gera a mensagem de PARABÉNS do email com o novo template e fonte"""
        
        nome_grupo = concessionaria

        mensagem = f"""
        <html>
        <head>
            <style>
                /* --- CORREÇÃO: 'font-weight: bold;' removido do body --- */
                body {{ 
                    font-family: Aptos, Arial, sans-serif;
                    font-size: 12pt;
                    font-weight: normal; /* <-- CORRIGIDO */
                }}
                /* --- CORREÇÃO: 'font-weight: bold;' adicionado ao .sucesso --- */
                .sucesso {{ 
                    color: #388E3C; 
                    font-weight: bold; /* <-- CORRIGIDO */
                }}
            </style>
        </head>
        <body>
            <p>Boa tarde a todos!</p>
            <p>Espero que estejam bem.</p>
            
            <p>Segue abaixo o acompanhamento do GMB do Grupo <strong>{nome_grupo}</strong>:</p>
            
            <p class="sucesso">Gostaríamos de agradecer ao Grupo
            {nome_grupo} pelo esforço e dedicação empenhados para enquadrar todas as suas
            unidades no objetivo estabelecido!</p>
            
            <p>No entanto, gostaríamos de
            reforçar que, embora o grupo esteja atualmente enquadrado, é fundamental
            continuar seguindo as dicas disponíveis no portal Reputation e manter o
            compromisso com as boas práticas recomendadas. Esses passos são essenciais
            para garantir que a pontuação se mantenha estável e não haja risco de queda.</p>

            <hr>
            <p style="font-style: italic;">Detalhes da apuração:</p>
            <p><img src="cid:{cid_imagem}"></p>

            <p>Mais uma vez, agradecemos o grupo pelos resultados
            alcançados até aqui e contamos com a continuidade desse resultado!!</p>
            
        </body>
        </html>
        """
        return mensagem
    
    def enviar_email(self, destinatarios_para, destinatarios_cc, assunto, corpo_html, anexo_inline_path=None, anexo_cid=None):
        """
        Exibe o email para teste com imagem inline (screenshot) E 
        a assinatura HTML personalizada com sua própria imagem.
        """
        try:
            email = self.outlook.CreateItem(0)
            
            email.To = destinatarios_para
            
            if destinatarios_cc:
                email.CC = destinatarios_cc
            
            email.Subject = assunto

            # --- 1. GERAR E ANEXAR ASSINATURA ---
            cid_imagem_assinatura = "img_assinatura_personalizada"
            assinatura_html = self._gerar_html_assinatura(cid_imagem_assinatura)
            
            caminho_imagem_assinatura = os.path.abspath(self.arquivo_imagem_assinatura)
            if os.path.exists(caminho_imagem_assinatura):
                attachment_sig = email.Attachments.Add(caminho_imagem_assinatura)
                PR_ATTACH_CONTENT_ID = "http://schemas.microsoft.com/mapi/proptag/0x3712001F"
                attachment_sig.PropertyAccessor.SetProperty(PR_ATTACH_CONTENT_ID, cid_imagem_assinatura)
            else:
                print(f"  ⚠ AVISO: Imagem da assinatura '{self.arquivo_imagem_assinatura}' não encontrada.")
            
            # --- 2. ANEXAR O SCREENSHOT (Lógica existente) ---
            if anexo_inline_path and anexo_cid and os.path.exists(anexo_inline_path):
                attachment_ss = email.Attachments.Add(os.path.abspath(anexo_inline_path))
                PR_ATTACH_CONTENT_ID = "http://schemas.microsoft.com/mapi/proptag/0x3712001F"
                attachment_ss.PropertyAccessor.SetProperty(PR_ATTACH_CONTENT_ID, anexo_cid)

            # --- 3. COMBINA MENSAGEM + ASSINATURA ---
            email.HTMLBody = corpo_html + assinatura_html

            # --- MODO DE TESTE ---
            email.Display() # Abre o email na tela para revisão
            
            print(f"  ✓ Email EXIBIDO PARA TESTE") 
            print(f"    Para: {destinatarios_para}")
            if destinatarios_cc:
                print(f"    Cc: {destinatarios_cc}")
            return True
            
        except Exception as e:
            print(f"  ✗ Erro ao exibir email: {e}")
            return False

    def _encontrar_colunas_datas(self, ws, linha_datas, max_col):
        """
        Lê a linha_datas, encontra todas as datas, e retorna o 
        índice da coluna mais recente e da segunda mais recente.
        """
        datas_encontradas = [] # (datetime_obj, col_index)
        
        for c in range(3, max_col + 1): 
            cell_val = ws.cell(linha_datas, c).value
            if isinstance(cell_val, datetime):
                datas_encontradas.append((cell_val, c))
        
        if not datas_encontradas:
            return None, None
            
        datas_encontradas.sort(key=lambda x: x[0], reverse=True)
        
        col_recente = datas_encontradas[0][1]
        col_anterior = datas_encontradas[1][1] if len(datas_encontradas) > 1 else None
        
        if col_anterior:
            print(f"    Debug Datas: Recente Col {col_recente}, Anterior Col {col_anterior}")
        else:
            print(f"    Debug Datas: Recente Col {col_recente} (Sem data anterior)")
            
        return col_recente, col_anterior

    # --- NOVO MÉTODO (MOVIDO PARA DENTRO DA CLASSE) ---
    def _identificar_subgrupos_regence_roma(self, nome_grupo, concessionarias_desenquadradas):
        """
        Separa as concessionárias desenquadradas de REGENCE e ROMA
        em subgrupos mapeados explicitamente para envio de emails.
        (Versão corrigida, sem o prefixo "Grupo ")
        """
        subgrupos = {}

        nome_grupo_upper = nome_grupo.upper()

        if "REGENCE" in nome_grupo_upper:
            regence_fortaleza = [c for c in concessionarias_desenquadradas if "FORTALEZA" in c.upper()]
            regence_veiculos = [c for c in concessionarias_desenquadradas if c not in regence_fortaleza]

            # --- CORREÇÃO AQUI ---
            if regence_fortaleza:
                subgrupos["Regence Veículos Fortaleza"] = regence_fortaleza
            if regence_veiculos:
                subgrupos["Regence Veículos"] = regence_veiculos
            # --- FIM DA CORREÇÃO ---

        elif "ROMA" in nome_grupo_upper:
            roma_mg = ["Roma Contagem", "Roma Norte"]
            roma_rj = ["Roma France Nova Iguaçu", "Roma France Recreio"]

            mg = [c for c in concessionarias_desenquadradas if c in roma_mg]
            rj = [c for c in concessionarias_desenquadradas if c in roma_rj]

            # --- CORREÇÃO AQUI ---
            if mg:
                subgrupos["Roma MG"] = mg
            if rj:
                subgrupos["Roma RJ"] = rj
            # --- FIM DA CORREÇÃO ---

        return subgrupos
    
    # --- MÉTODO ATUALIZADO (MOVIDO PARA DENTRO DA CLASSE) ---
    def analisar_e_processar_erep(self, aba_dados_tabelas, linha_inicial_scan, limite_pontuacao):
        """
        Varre a Coluna B procurando por grupos "(E-REP)".
        Lê os nomes das concessionárias da Coluna H.
        Analisa as notas e envia emails (Alerta ou Parabéns).
        Inclui regras especiais para ROMA e REGENCE e ignora linhas de nota geral.
        """
    
        destinatarios_df = self.carregar_destinatarios()
        if destinatarios_df is None:
            return 
            
        print(f"✓ Destinatários carregados de '{self.arquivo_destinatarios}'")

        pasta_temp = "screenshots_temp"
        if not os.path.exists(pasta_temp):
            os.makedirs(pasta_temp)

        try:
            print(f"Abrindo arquivo principal '{self.arquivo_dados}' para análise...")
            wb = openpyxl.load_workbook(self.arquivo_dados, data_only=True)
            ws = wb[aba_dados_tabelas]
        except Exception as e:
            print(f"✗ ERRO AO ABRIR '{self.arquivo_dados}' (Aba: {aba_dados_tabelas}): {e}")
            return

        print(f"\n{'='*60}")
        print(f"INICIANDO ANÁLISE DA ABA '{aba_dados_tabelas}' (MODO DE TESTE)")
        print(f"Gatilho do Bloco: Célula com '(E-REP)' na Coluna B")
        print(f"Nome da Concessionária: Lendo da Coluna H")
        print(f"Lógica: Análise de colunas de data dinâmicas.")
        print(f"Limite de Pontuação: < {limite_pontuacao}")
        print(f"{'='*60}\n")
        
        linha_atual = linha_inicial_scan
        max_linha = ws.max_row
        max_col_planilha = ws.max_column
        
        COLUNA_GATILHO_GRUPO = 2  # Coluna B
        COLUNA_NOME_CONCESSIONARIA = 8  # Coluna H
        
        grupos_enviados = 0
        data_hoje_formatada = datetime.now().strftime("%d.%m")

        while linha_atual <= max_linha:
            cell_value = ws.cell(linha_atual, COLUNA_GATILHO_GRUPO).value
            
            if cell_value and "(E-REP)" in str(cell_value):
                nome_grupo = str(cell_value).replace("(E-REP)", "").strip()
                linha_inicio_bloco = linha_atual
                linha_datas = linha_atual + 1
                linha_inicio_dados = linha_atual + 2
                
                print(f"--- Processando Bloco: {nome_grupo} (Início Linha {linha_inicio_bloco}) ---")

                # --- Encontrar o fim do bloco ---
                linha_fim_bloco = linha_inicio_dados
                while linha_fim_bloco <= (max_linha + 3):
                    try:
                        proxima_celula = ws.cell(linha_fim_bloco, COLUNA_GATILHO_GRUPO).value
                        if proxima_celula and "(E-REP)" in str(proxima_celula):
                            break
                        
                        linha_vazia = all(ws.cell(linha_fim_bloco + i, c).value is None for i in range(3) for c in range(1, 20))
                        if linha_vazia:
                            break
                    except IndexError:
                        break
                    linha_fim_bloco += 1
                
                linha_fim_bloco = min(linha_fim_bloco - 1, max_linha)
                print(f"    Debug Bloco: Fim na linha {linha_fim_bloco}.")
                
                # --- Encontrar colunas de data ---
                col_recente, col_anterior = self._encontrar_colunas_datas(ws, linha_datas, max_col_planilha)
                if col_recente is None:
                    print(f"  ⚠ AVISO: Não foi possível encontrar datas válidas para {nome_grupo}. Pulando.")
                    print("-" * 30)
                    linha_atual = linha_fim_bloco + 1
                    continue

                flag_alerta_recente = False
                flag_alerta_anterior = False
                concessionarias_desenquadradas = []

                # --- Analisar cada linha do grupo ---
                for r in range(linha_inicio_dados, linha_fim_bloco + 1):
                    nome_sub = ws.cell(r, COLUNA_NOME_CONCESSIONARIA).value
                    score_recente = ws.cell(r, col_recente).value

                    if nome_sub is None or str(nome_sub).strip() == "":
                        continue

                    nome_upper = str(nome_sub).upper()

                    # Ignorar linhas de "nota geral"
                    if any(palavra in nome_upper for palavra in ["TOTAL", "GERAL", "MÉDIA", "MEDIA", "RESULTADO"]):
                        continue

                    # Verificar pontuação abaixo do limite
                    if isinstance(score_recente, (int, float)) and score_recente < limite_pontuacao:
                        flag_alerta_recente = True
                        nome_sub_limpo = str(nome_sub).strip()
                        
                        concessionarias_desenquadradas.append(nome_sub_limpo)

                    # Verifica data anterior (para detectar melhora)
                    if col_anterior:
                        score_ant = ws.cell(r, col_anterior).value
                        if isinstance(score_ant, (int, float)) and score_ant < limite_pontuacao:
                            flag_alerta_anterior = True

                # --- Remove duplicados ---
                concessionarias_desenquadradas = list(dict.fromkeys(concessionarias_desenquadradas))

                # --- TRATAMENTO ESPECIAL PARA REGENCE E ROMA ---
                if ("REGENCE" in nome_grupo.upper() or "ROMA" in nome_grupo.upper()) and flag_alerta_recente:
                    subgrupos = self._identificar_subgrupos_regence_roma(nome_grupo, concessionarias_desenquadradas)
                    
                    if not subgrupos:
                        print(f"  ⚠ Grupo {nome_grupo}: Sem concessionárias válidas para envio.")
                        linha_atual = linha_fim_bloco + 1
                        continue
                    
                    caminho_img = os.path.join(pasta_temp, f"{nome_grupo.replace(' ', '_').replace('/', '-')}.png")
                    col_fim = 1
                    for r_ss in range(linha_inicio_bloco, linha_fim_bloco + 1):
                        for c in range(max_col_planilha, 1, -1):
                            if ws.cell(r_ss, c).value is not None and c > col_fim:
                                col_fim = c
                                break

                    screenshot = self.tirar_screenshot_tabela(
                        self.arquivo_dados,
                        aba_dados_tabelas,
                        linha_inicio_bloco,
                        linha_fim_bloco,
                        COLUNA_GATILHO_GRUPO,
                        col_fim,
                        caminho_img
                    )
                    
                    # Envia um email para cada subgrupo (usando o nome limpo)
                    for nome_subgrupo, lista_desenquadradas_sub in subgrupos.items():
                        # 'nome_subgrupo' agora é "Roma MG", que vai bater com a lista
                        dest = destinatarios_df[destinatarios_df['Grupo'] == nome_subgrupo]
                        
                        if dest.empty:
                            print(f"  ⚠ Nenhum destinatário encontrado para '{nome_subgrupo}'. Pulando.")
                            continue
                        
                        emails_para = dest['Para'].values[0]
                        emails_cc = dest['Cc'].values[0] if 'Cc' in dest.columns and pd.notna(dest['Cc'].values[0]) else ''
                        
                        cid_img = f"relatorio_img_{nome_subgrupo.replace(' ', '_')}.png"
                        corpo = self.gerar_mensagem_alerta(nome_subgrupo, lista_desenquadradas_sub, cid_img)
                        assunto = f"Acompanhamento GMB {nome_subgrupo} | {data_hoje_formatada}"
                        
                        print(f"  ⚠ Subgrupo {nome_subgrupo}: Pontuação abaixo do limite — Gerando ALERTA.")
                        
                        self.enviar_email(
                            emails_para,
                            emails_cc,
                            assunto,
                            corpo,
                            anexo_inline_path=screenshot,
                            anexo_cid=cid_img
                        )
                        grupos_enviados += 1
                
                # --- TRATAMENTO PARA PARABÉNS (REGENCE E ROMA) ---
                elif ("REGENCE" in nome_grupo.upper() or "ROMA" in nome_grupo.upper()) and not flag_alerta_recente and flag_alerta_anterior:
                    subgrupos_names = []
                    
                    # --- CORREÇÃO AQUI: Nomes sem o prefixo "Grupo " ---
                    if "REGENCE" in nome_grupo.upper():
                        subgrupos_names = ["Regence Veículos", "Regence Veículos Fortaleza"]
                    elif "ROMA" in nome_grupo.upper():
                        subgrupos_names = ["Roma MG", "Roma RJ"]
                    # --- FIM DA CORREÇÃO ---
                    
                    caminho_img = os.path.join(pasta_temp, f"{nome_grupo.replace(' ', '_').replace('/', '-')}.png")
                    col_fim = 1
                    for r_ss in range(linha_inicio_bloco, linha_fim_bloco + 1):
                        for c in range(max_col_planilha, 1, -1):
                            if ws.cell(r_ss, c).value is not None and c > col_fim:
                                col_fim = c
                                break

                    screenshot = self.tirar_screenshot_tabela(
                        self.arquivo_dados,
                        aba_dados_tabelas,
                        linha_inicio_bloco,
                        linha_fim_bloco,
                        COLUNA_GATILHO_GRUPO,
                        col_fim,
                        caminho_img
                    )
                    
                    for nome_subgrupo in subgrupos_names:
                        dest = destinatarios_df[destinatarios_df['Grupo'] == nome_subgrupo]
                        
                        if dest.empty:
                            print(f"  ⚠ Nenhum destinatário encontrado para '{nome_subgrupo}'. Pulando.")
                            continue
                        
                        emails_para = dest['Para'].values[0]
                        emails_cc = dest['Cc'].values[0] if 'Cc' in dest.columns and pd.notna(dest['Cc'].values[0]) else ''
                        
                        cid_img = f"relatorio_img_{nome_subgrupo.replace(' ', '_')}.png"
                        corpo = self.gerar_mensagem_parabens(nome_subgrupo, cid_img)
                        assunto = f"Acompanhamento GMB {nome_subgrupo} | {data_hoje_formatada}"
                        
                        print(f"  ✓ Subgrupo {nome_subgrupo}: Recuperação detectada — Gerando PARABÉNS.")
                        
                        self.enviar_email(
                            emails_para,
                            emails_cc,
                            assunto,
                            corpo,
                            anexo_inline_path=screenshot,
                            anexo_cid=cid_img
                        )
                        grupos_enviados += 1
                
                # --- TRATAMENTO NORMAL PARA OUTROS GRUPO (Já estava correto) ---
                else:
                    dest = destinatarios_df[destinatarios_df['Grupo'] == nome_grupo]
                    if dest.empty:
                        print(f"  ⚠ Nenhum destinatário encontrado para '{nome_grupo}'. Pulando.\n")
                        linha_atual = linha_fim_bloco + 1
                        continue

                    emails_para = dest['Para'].values[0]
                    emails_cc = dest['Cc'].values[0] if 'Cc' in dest.columns and pd.notna(dest['Cc'].values[0]) else ''

                    corpo = None
                    assunto = None
                    cid_img = f"relatorio_img_{nome_grupo.replace(' ', '_')}.png"

                    if flag_alerta_recente:
                        print(f"  ⚠ Grupo {nome_grupo}: Pontuação abaixo do limite — Gerando ALERTA.")
                        corpo = self.gerar_mensagem_alerta(nome_grupo, concessionarias_desenquadradas, cid_img)
                        assunto = f"Acompanhamento GMB {nome_grupo} | {data_hoje_formatada}"
                        grupos_enviados += 1

                    elif not flag_alerta_recente and flag_alerta_anterior:
                        print(f"  ✓ Grupo {nome_grupo}: Recuperação detectada — Gerando PARABÉNS.")
                        corpo = self.gerar_mensagem_parabens(nome_grupo, cid_img)
                        assunto = f"Acompanhamento GMB {nome_grupo} | {data_hoje_formatada}"
                        grupos_enviados += 1

                    else:
                        print(f"  - Grupo {nome_grupo}: Tudo OK. Nenhum email enviado.")

                    if corpo and assunto:
                        caminho_img = os.path.join(pasta_temp, f"{nome_grupo.replace(' ', '_').replace('/', '-')}.png")

                        col_fim = 1
                        for r_ss in range(linha_inicio_bloco, linha_fim_bloco + 1):
                            for c in range(max_col_planilha, 1, -1):
                                if ws.cell(r_ss, c).value is not None and c > col_fim:
                                    col_fim = c
                                    break

                        screenshot = self.tirar_screenshot_tabela(
                            self.arquivo_dados,
                            aba_dados_tabelas,
                            linha_inicio_bloco,
                            linha_fim_bloco,
                            COLUNA_GATILHO_GRUPO,
                            col_fim,
                            caminho_img
                        )

                        self.enviar_email(
                            emails_para,
                            emails_cc,
                            assunto,
                            corpo,
                            anexo_inline_path=screenshot,
                            anexo_cid=cid_img
                        )

                print("-" * 30)
                linha_atual = linha_fim_bloco + 1

            else:
                linha_atual += 1

        wb.close()
        print(f"\n{'='*60}")
        print(f"PROCESSO CONCLUÍDO (MODO DE TESTE)")
        print(f"Total de emails gerados: {grupos_enviados}")
        print(f"{'='*60}\n")

# -----------------------------------------------------------------
# BLOCO DE EXECUÇÃO - CONFIGURAÇÃO FINAL
# -----------------------------------------------------------------
if __name__ == "__main__":
    # Configurações
    
    # Arquivo com os emails (Para/Cc) e Concessionária
    ARQUIVO_DESTINATARIOS = "Emails grupo.xlsm"
    ABA_DESTINATARIOS = "BaseEnvios"
    
    # Arquivo que contém TUDO (pontuações E tabelas)
    ARQUIVO_PRINCIPAL = "Acompanhamento dispersão.xlsx" 
    
    # Aba onde TUDO será analisado
    ABA_E_REP = "E-REP"
    
    # Linha onde o script deve começar a procurar pelos grupos
    LINHA_INICIAL_SCAN = 30
    
    LIMITE_PONTUACAO = 780
    
    # --- Executa ---
    print("Iniciando automação (Modo E-REP / Lógica de Datas Dinâmicas)...")
    print(f"Caminho atual: {os.getcwd()}")
    
    if not os.path.exists(ARQUIVO_DESTINATARIOS):
        print(f"ERRO: Arquivo de destinatários NÃO ENCONTRADO: {ARQUIVO_DESTINATARIOS}")
    elif not os.path.exists(ARQUIVO_PRINCIPAL):
        print(f"ERRO: Arquivo principal NÃO ENCONTRADO: {ARQUIVO_PRINCIPAL}")
    else:
        automacao = AutomacaoEmailOutlook(
            ARQUIVO_DESTINATARIOS, 
            ARQUIVO_PRINCIPAL, 
            ABA_DESTINATARIOS
        )
        
        # Chama a função de processamento único
        automacao.analisar_e_processar_erep(
            aba_dados_tabelas=ABA_E_REP,
            linha_inicial_scan=LINHA_INICIAL_SCAN,
            limite_pontuacao=LIMITE_PONTUACAO
        )